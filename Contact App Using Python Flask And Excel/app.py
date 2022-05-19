from flask import Flask, render_template, request, flash
import logging
import os
from openpyxl import load_workbook
app = Flask(__name__)
app.secret_key = 'MyKey'
logging.basicConfig(filename='.//debug.log', level=logging.DEBUG,
                    format=f'%(asctime)s %(levelname)s %(name)s %(threadName)s : %(message)s')


@app.route('/', methods=['GET', 'POST'])
def MyApp():
    if request.method == 'POST':
        name = request.form['name']
        mno = request.form['mno']
        myFileName = './/Contacts.xlsx'

        try:
            wb = load_workbook(filename=myFileName)
            ws = wb['Sheet1']
            newRowLocation = ws.max_row + 1
            ws.cell(column=2, row=newRowLocation, value=name)
            ws.cell(column=3, row=newRowLocation, value=mno)
            wb.save(filename=myFileName)
            wb.close()

            flash("Contact has been saved.")
            return render_template('index.html', **locals())
        except Exception as e:

            flash("An Error Occured")
            return render_template('index.html', **locals())
    else:
        return render_template('index.html', **locals())


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
